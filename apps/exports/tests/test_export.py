from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from apps.banking.models import Account, Institution, Transaction
from apps.investments.models import InvestmentAccount, Holding
from apps.assets.models import Asset
from apps.liabilities.models import Liability


pytestmark = pytest.mark.django_db


@pytest.fixture
def alice(django_user_model):
    return django_user_model.objects.create_user(username="alice", password="x")


@pytest.fixture
def alice_client(alice):
    c = Client()
    c.force_login(alice)
    return c


def _load(response):
    return load_workbook(BytesIO(response.content))


def test_export_returns_xlsx(alice_client):
    response = alice_client.get(reverse("exports:xlsx"))
    assert response.status_code == 200
    assert response["Content-Type"].startswith("application/vnd.openxmlformats")
    assert "attachment" in response["Content-Disposition"]


def test_export_contains_bank_account_sheet(alice, alice_client):
    inst = Institution.objects.create(user=alice, name="Bank", access_url="https://x")
    acc = Account.objects.create(institution=inst, name="Checking", type="checking",
                                 balance=Decimal("100"), external_id="A-1")
    Transaction.objects.create(account=acc, posted_at=date(2026, 4, 1),
                               amount=Decimal("-50"), payee="Coffee", external_id="t-1")
    response = alice_client.get(reverse("exports:xlsx"))
    wb = _load(response)
    assert "Checking" in wb.sheetnames
    ws = wb["Checking"]
    # Header + 1 row
    assert ws.max_row == 2
    assert ws.cell(row=2, column=2).value == "Coffee"


def test_export_includes_holdings_assets_liabilities_sheets(alice, alice_client):
    response = alice_client.get(reverse("exports:xlsx"))
    wb = _load(response)
    for name in ("Holdings", "Assets", "Liabilities"):
        assert name in wb.sheetnames


def test_export_excludes_other_users_data(alice, alice_client, django_user_model):
    bob = django_user_model.objects.create_user(username="bob", password="x")
    bob_inst = Institution.objects.create(user=bob, name="Bob Bank", access_url="https://b")
    Account.objects.create(institution=bob_inst, name="BobAccount", type="checking",
                           balance=Decimal("0"), external_id="B-1")
    response = alice_client.get(reverse("exports:xlsx"))
    wb = _load(response)
    assert "BobAccount" not in wb.sheetnames


def test_export_requires_login():
    c = Client()
    response = c.get(reverse("exports:xlsx"))
    assert response.status_code == 302
    assert "/login/" in response["Location"]
