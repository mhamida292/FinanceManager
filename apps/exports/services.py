from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.assets.models import Asset
from apps.banking.models import Account
from apps.investments.models import InvestmentAccount
from apps.liabilities.models import Liability


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="404040")
RIGHT = Alignment(horizontal="right")
MONEY_FORMAT = '"$"#,##0.00;[Red]-"$"#,##0.00'


def _autosize(ws, min_width: int = 10, max_width: int = 50) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for cell in ws[letter]:
            longest = max(longest, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[letter].width = max(min_width, min(longest + 2, max_width))


def _write_header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def build_workbook(*, user) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ----- Bank accounts: one sheet per account, sheet contains transactions
    accounts = (
        Account.objects.for_user(user)
        .select_related("institution")
        .order_by("institution__name", "name")
    )
    for acc in accounts:
        title = (acc.effective_name or f"Account {acc.id}")[:31]  # excel sheet titles max 31 chars
        ws = wb.create_sheet(title=title)
        _write_header(ws, ["Date", "Payee", "Memo", "Amount", "Pending"])
        rows = acc.transactions.order_by("-posted_at", "-id")
        for tx in rows:
            ws.append([
                tx.posted_at,
                tx.payee or tx.description or "",
                tx.memo or "",
                float(tx.amount),
                "yes" if tx.pending else "",
            ])
        # Format amount column as currency
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=4).number_format = MONEY_FORMAT
            ws.cell(row=r, column=4).alignment = RIGHT
        _autosize(ws)

    # ----- Holdings sheet (all investment accounts in one sheet)
    ws = wb.create_sheet(title="Holdings")
    _write_header(ws, ["Account", "Broker", "Symbol", "Shares", "Price", "Market value", "Cost basis", "Gain $", "Gain %"])
    inv_accounts = InvestmentAccount.objects.for_user(user).prefetch_related("holdings").order_by("broker", "name")
    for acc in inv_accounts:
        for h in acc.holdings.all().order_by("symbol"):
            ws.append([
                acc.effective_name,
                acc.broker or "",
                h.symbol,
                float(h.shares),
                float(h.current_price or 0),
                float(h.market_value or 0),
                float(h.cost_basis) if h.cost_basis is not None else None,
                float(h.gain_loss) if h.gain_loss is not None else None,
                float(h.gain_loss_percent) if h.gain_loss_percent is not None else None,
            ])
    for r in range(2, ws.max_row + 1):
        for col in (5, 6, 7, 8):  # Price, Value, Cost, Gain $
            ws.cell(row=r, column=col).number_format = MONEY_FORMAT
            ws.cell(row=r, column=col).alignment = RIGHT
        ws.cell(row=r, column=9).number_format = "0.00"  # Gain %
    _autosize(ws)

    # ----- Assets sheet
    ws = wb.create_sheet(title="Assets")
    _write_header(ws, ["Name", "Kind", "Quantity", "Unit", "Value", "Last priced", "Notes"])
    for a in Asset.objects.for_user(user).order_by("name"):
        ws.append([
            a.name,
            a.kind,
            float(a.quantity) if a.quantity else "",
            a.unit or "",
            float(a.current_value or 0),
            a.last_priced_at,
            a.notes or "",
        ])
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = MONEY_FORMAT
        ws.cell(row=r, column=5).alignment = RIGHT
    _autosize(ws)

    # ----- Liabilities sheet
    ws = wb.create_sheet(title="Liabilities")
    _write_header(ws, ["Name", "Balance owed", "Notes"])
    for lia in Liability.objects.for_user(user).order_by("name"):
        ws.append([lia.name, float(lia.balance), lia.notes or ""])
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).number_format = MONEY_FORMAT
        ws.cell(row=r, column=2).alignment = RIGHT
    _autosize(ws)

    return wb
